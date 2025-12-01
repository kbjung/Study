#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
pepbind_pipeline.py - WSL/오프라인 환경용 통합 파이프라인 (정리 버전)

구성:
- STEP 2: PepMLM(ESM-2)로 펩타이드 후보 생성 (GPU 사용)
- STEP 3: ColabFold 멀티머로 타깃-펩타이드 복합체 구조 예측 (진행 상황 표시)
- STEP 4: AutoDock Vina 도킹 (CPU, stdout 파싱)
- STEP 5: PLIP 상호작용 분석
- STEP 6: PRODIGY 결합 자유에너지 평가
- STEP 7: 최종 평가(A안 가중치) + 엑셀 파일 생성 + rank_001 PDB zip 압축

A안 가중치:
  PRODIGY 0.50
  Vina    0.25
  PLIP    0.15
  ipTM    0.10
"""

import os
import time
import csv
import re
import json
import zipfile
import shutil
import subprocess
import gc
from pathlib import Path
from datetime import datetime

import torch
import torch.nn.functional as F
from transformers import AutoTokenizer, AutoModelForMaskedLM
from Bio.PDB import PDBParser

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
import numpy as np

import xml.etree.ElementTree as ET
from collections import defaultdict
import math
import sys


START_TIME = datetime.now()
END_TIME = None          # 전체 종료시간 저장용
STEP_TIMINGS = []        # 각 스텝별 시작/종료/소요시간 기록용

# =====================================================================
# === 사용자 설정 영역: 여기만 수정해서 사용 ==========================
# =====================================================================

# 1) 타깃 단백질 서열 (FASTA의 sequence 부분만)
TARGET_SEQUENCE = (
    "AFTVTVPKDLYVVEYGSNMTIECKFPVEKQLDLAALIVYWEMEDKNIIQFVHGEEDLKVQHSSYRQRARLLKDQLSLGNAALQITDVKLQDAGVYRCMISYGGADYKRITVKVNAPYNKINQRILVVDPVTSEHELTCQAEGYPKAEVIWTSSDHQVLSGKTTTTNSKREEKLFNVTSTLRINTTTNEIFYCTFRRLDPEENHTAELVIPELPLAHPPNERT" # PD-L1 단백질 서열
)

# 2) 생성할 펩타이드 설정
NUM_PEPTIDES   = 10   # 생성할 펩타이드 후보 개수
PEPTIDE_LENGTH = 10    # 각 펩타이드 길이 (아미노산 개수)

# 3) ColabFold / 평가 단계 사용 여부
RUN_COLABFOLD  = True   # ColabFold 구조 예측 실행 여부
RUN_VINA       = True   # AutoDock Vina 도킹 실행 여부
RUN_PLIP       = True   # PLIP 상호작용 분석 실행 여부
RUN_PRODIGY    = True   # PRODIGY 결합 친화도 평가 실행 여부

# 4) 작업 기본 디렉토리
BASE_DIR = Path(os.environ.get("PEPBIND_BASE_DIR", "~/work/pipeline")).expanduser()

# 5) 외부 도구 경로 (환경에 맞게 수정 가능)
COLABFOLD_CMD   = os.environ.get("COLABFOLD_CMD", "colabfold_batch").strip()
VINA_CMD        = os.environ.get("VINA_CMD", "vina").strip()
PLIP_CMD        = os.environ.get("PLIP_CMD", "plip").strip()          # 기본값도 plip으로
PRODIGY_SCRIPT  = os.environ.get("PRODIGY_SCRIPT", "prodigy").strip()

OBABEL_CMD = shutil.which("obabel") or "obabel"

# ColabFold 자원/안전 관련 설정
# 기본값은 32:128, 메모리 많이 부족하면 환경변수나 여기 값을 "16:64"로 줄여도 됨
COLABFOLD_MAX_MSA = os.environ.get("COLABFOLD_MAX_MSA", "16:64")

# 진행률이 일정 시간 이상 변화 없으면 강제 종료 (메모리 부족/프리징 방지용)
COLABFOLD_MAX_IDLE_MIN = int(os.environ.get("COLABFOLD_MAX_IDLE_MIN", "30"))   # 예: 30분

# 전체 ColabFold 실행 시간 상한 (분)
COLABFOLD_MAX_TOTAL_MIN = int(os.environ.get("COLABFOLD_MAX_TOTAL_MIN", "1440"))  # 예: 360(6시간)

# GPU에서 OOM(RESOURCE_EXHAUSTED / Out of memory) 나면 CPU로 한 번 더 재시도할지 여부
COLABFOLD_CPU_FALLBACK = os.environ.get("COLABFOLD_CPU_FALLBACK", "1").lower() not in (
    "0", "false", "no", "off"
)

# PepMLM 하이퍼파라미터 설정
PEPMLM_TOP_K = 10 # 예. 후보군 Top10 : 10
PEPMLM_TEMPERATURE = 1.0 # default 1.0

# =====================================================================
# === 공통 설정 / 유틸 =================================================
# =====================================================================

BASE_DIR.mkdir(parents=True, exist_ok=True)

DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
print(f"[INFO] PyTorch device: {DEVICE}")

# JAX / ColabFold 메모리 설정 (모든 자식 프로세스에 적용)
os.environ["XLA_PYTHON_CLIENT_PREALLOCATE"] = "false"
os.environ["XLA_PYTHON_CLIENT_MEM_FRACTION"] = "0.8"

# 요약/최종 테이블 컬럼 설정 (여기만 고치면 순서가 반영되도록)
VINA_SUMMARY_COLS = [
    "complex",
    "vina_status",
    "vina_score",
    "receptor_pdbqt",
    "ligand_pdbqt",
    "log_file",
]

PLIP_SUMMARY_COLS = [
    "complex",
    "plip_status",
    "plip_total_interactions",
    "plip_hbond",
    "plip_hydrophobic",
    "plip_saltbridge",
]

PRODIGY_SUMMARY_COLS = [
    "complex",
    "PRODIGY_status",
    "PRODIGY_dG",
]

FINAL_TABLE_HEADERS = [
    "rank",
    "candidate_id",
    "peptide_seq",
    "complex_pdb",
    "AlphaFold_status",
    "FinalScore",
    "PRODIGY_status",
    "PRODIGY_dG(kcal/mol)",
    "Vina_status",
    "Vina_score(kcal/mol)",
    "PLIP_status",
    "PLIP_total_interactions",
    "PLIP_hbond",
    "PLIP_hydrophobic",
    "PLIP_saltbridge",
    "ipTM",
]


def timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def format_elapsed(start: datetime, end: datetime) -> str:
    """두 시각 차이를 '00일 00시간 00분 00초' 형식 문자열로 변환."""
    elapsed = end - start
    total_seconds = int(elapsed.total_seconds())

    days = total_seconds // (24 * 3600)
    total_seconds %= (24 * 3600)
    hours = total_seconds // 3600
    total_seconds %= 3600
    minutes = total_seconds // 60
    seconds = total_seconds % 60

    parts = []
    if days > 0:
        parts.append(f"{days:02d}일")
    if days > 0 or hours > 0:
        parts.append(f"{hours:02d}시간")
    parts.append(f"{minutes:02d}분")
    parts.append(f"{seconds:02d}초")
    return " ".join(parts)


def format_seconds_hms(seconds: float) -> str:
    """
    초 단위 시간 값을 '00시간 00분 00초' 형식 문자열로 변환.
    (1개 샘플당 평균 소요 시간 표시용)
    """
    if seconds < 0:
        seconds = 0

    total_seconds = int(round(seconds))

    hours = total_seconds // 3600
    total_seconds %= 3600
    minutes = total_seconds // 60
    secs = total_seconds % 60

    parts = []
    if hours > 0:
        parts.append(f"{hours:02d}시간")
    if hours > 0 or minutes > 0:
        parts.append(f"{minutes:02d}분")
    parts.append(f"{secs:02d}초")

    return " ".join(parts)


def print_step_timing(step_label: str, start: datetime, end: datetime):
    """각 스텝의 시작/종료/소요 시간을 출력 + 전역 리스트에 저장."""
    elapsed_str = format_elapsed(start, end)

    print("\n" + "-" * 80)
    print(f"[STEP TIMER] {step_label}")
    print(f"  시작: {start.strftime('%Y.%m.%d %H:%M:%S')}")
    print(f"  종료: {end.strftime('%Y.%m.%d %H:%M:%S')}")
    print(f"  소요 시간: {elapsed_str}")
    print("-" * 80)

    # 전역 리스트에 기록
    global STEP_TIMINGS
    STEP_TIMINGS.append(
        {
            "step": step_label,
            "start": start,
            "end": end,
            "elapsed": elapsed_str,
        }
    )


def init_workspace():
    """PDP_YYYYMMDD_HHMMSS 형태 워크스페이스 및 하위 폴더 생성."""
    ws_name = f"PDP_{timestamp()}"
    ws_root = BASE_DIR / ws_name

    folders = {
        "root": ws_root,
        "fasta": ws_root / "fasta",
        "pdb": ws_root / "pdb",
        "colabfold_out": ws_root / "pdb" / "colabfold_output",
        "results": ws_root / "results",
        "vina": ws_root / "results" / "vina",
        "plip": ws_root / "results" / "plip",
        "prodigy": ws_root / "results" / "prodigy",
        "temp": ws_root / "temp",
    }
    for d in folders.values():
        d.mkdir(parents=True, exist_ok=True)

    print("\n" + "=" * 80)
    print("STEP 1: 워크스페이스 / 폴더 구조 생성")
    print("=" * 80)
    for k, v in folders.items():
        print(f"✔️ {k:12s}: {v}")
    print("=" * 80)

    return folders


def parse_prodigy_dg_from_stdout(stdout: str):
    """
    PRODIGY stdout에서 ΔG(또는 binding energy)를 추출하는 헬퍼 함수.
    형식 예시:
      Binding energy: -12.3 kcal/mol
      Predicted ΔG: -10.5 kcal/mol

    여러 줄 중 첫 번째 매칭값만 사용.
    실패하면 None 리턴.
    """
    if not stdout:
        return None

    # 부호로 허용할 문자: ASCII -, +, 그리고 유니코드 마이너스(−, U+2212)
    sign_pattern = r"[+\-−]?"

    patterns = [
        rf"Binding energy\s*[:=]\s*({sign_pattern}\d+(?:\.\d+)?)",
        rf"Predicted\s*Δ?G\s*[:=]\s*({sign_pattern}\d+(?:\.\d+)?)",
        rf"\bΔG\s*[:=]\s*({sign_pattern}\d+(?:\.\d+)?)",
    ]

    for pat in patterns:
        m = re.search(pat, stdout, re.IGNORECASE)
        if m:
            raw = m.group(1)
            # 유니코드 마이너스(−)를 ASCII - 로 치환
            raw = raw.replace("−", "-")
            try:
                return float(raw)
            except ValueError:
                pass

    # 2) 백업: stdout 전체에서 "소수점이 있는 첫 번째 실수"
    m = re.search(rf"({sign_pattern}\d+\.\d+)", stdout)
    if m:
        raw = m.group(1).replace("−", "-")
        try:
            return float(raw)
        except ValueError:
            return None

    return None


def extract_chain_as_pdb(complex_pdb: Path, chain_id: str, out_pdb: Path):
    """
    complex_pdb에서 특정 chain_id(A/B 등) 전체를 골라
    하나의 PDB 파일로 저장.
    ATOM, HETATM만 대상.
    """
    lines_out = []
    with open(complex_pdb, "r") as f:
        for line in f:
            if not (line.startswith("ATOM") or line.startswith("HETATM")):
                continue
            if len(line) < 22:
                continue
            if line[21] == chain_id:
                lines_out.append(line.rstrip("\n"))

    if not lines_out:
        raise ValueError(f"{complex_pdb}에서 체인 {chain_id} 원자를 찾지 못했습니다.")

    with open(out_pdb, "w") as f:
        for ln in lines_out:
            f.write(ln + "\n")
        f.write(f"TER\n")
        f.write("END\n")


def clear_gpu_memory():
    """
    ColabFold 실행 전/후에 PyTorch CUDA 캐시와 파이썬 객체를 정리해서
    GPU 메모리를 최대한 비워주는 헬퍼 함수.
    (colabfold_batch 자체는 별도 프로세스라, 그쪽에서 쓰던 VRAM은
     프로세스 종료 시 자동으로 해제된다.)
    """
    try:
        if torch.cuda.is_available():
            print("[INFO] GPU 메모리 초기화: torch.cuda.empty_cache() 실행")
            torch.cuda.empty_cache()
            try:
                # 일부 환경에서 shared memory 핸들도 정리
                torch.cuda.ipc_collect()
            except Exception:
                pass
        else:
            print("[INFO] CUDA 장치가 없어 GPU 초기화는 건너뜁니다.")
    except Exception as e:
        print(f"[WARN] GPU 메모리 초기화 중 예외 발생: {e}")

    # 파이썬 객체 GC
    gc.collect()
    print("[INFO] Python GC 실행 완료")


class Tee:
    """
    sys.stdout / sys.stderr를 여러 스트림(터미널 + 로그 파일)에 동시에 보내기 위한 헬퍼.
    """
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data):
        for s in self.streams:
            s.write(data)
        for s in self.streams:
            s.flush()

    def flush(self):
        for s in self.streams:
            s.flush()


def setup_logging(log_path: Path):
    """
    전체 파이프라인 출력(sys.stdout, sys.stderr)을 터미널 + 로그 파일로 동시에 보내기.

    - log_path: 생성할 로그 파일 경로
    """
    # 'w' 로 새로 열고, utf-8로 저장
    log_file = open(log_path, "w", encoding="utf-8")
    tee = Tee(sys.stdout, log_file)
    sys.stdout = tee
    sys.stderr = tee

    print(f"[LOG] 전체 출력 로그 파일: {log_path}")


def print_run_config():
    """
    이번 pepbind 실행에 사용된 주요 옵션/환경을 한 번에 출력 (로그에 같이 기록됨).
    """
    print("\n" + "=" * 80)
    print("PEPBIND RUN CONFIG")
    print("=" * 80)
    print(f"TARGET_SEQUENCE length: {len(TARGET_SEQUENCE.strip())}")
    print(f"NUM_PEPTIDES          : {NUM_PEPTIDES}")
    print(f"PEPTIDE_LENGTH        : {PEPTIDE_LENGTH}")
    print(f"PepMLM_top_k          : {PEPMLM_TOP_K}")
    print(f"PepMLM_temperature    : {PEPMLM_TEMPERATURE}")
    print(f"RUN_COLABFOLD         : {RUN_COLABFOLD}")
    print(f"RUN_VINA              : {RUN_VINA}")
    print(f"RUN_PLIP              : {RUN_PLIP}")
    print(f"RUN_PRODIGY           : {RUN_PRODIGY}")
    print(f"BASE_DIR              : {BASE_DIR}")
    print(f"COLABFOLD_CMD         : {COLABFOLD_CMD}")
    print(f"VINA_CMD              : {VINA_CMD}")
    print(f"PLIP_CMD              : {PLIP_CMD}")
    print(f"PRODIGY_SCRIPT        : {PRODIGY_SCRIPT}")
    print(f"OBABEL_CMD            : {OBABEL_CMD}")
    print(f"COLABFOLD_MAX_MSA     : {COLABFOLD_MAX_MSA}")
    print(f"COLABFOLD_MAX_IDLE_MIN: {COLABFOLD_MAX_IDLE_MIN}")
    print(f"COLABFOLD_MAX_TOTAL_MIN: {COLABFOLD_MAX_TOTAL_MIN}")
    print(f"COLABFOLD_CPU_FALLBACK: {COLABFOLD_CPU_FALLBACK}")
    print(f"PyTorch DEVICE        : {DEVICE}")
    print("=" * 80 + "\n")


def autofit_worksheet_columns(ws):
    """
    워크시트의 각 열에 대해
    - 셀 값의 문자열 길이를 기준으로
    - 열 너비를 대략적으로 자동 조절해주는 함수.
    """
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            # 숫자, 날짜 등도 문자열 길이로 변환
            value_str = str(value)
            if len(value_str) > max_length:
                max_length = len(value_str)

        # 여유 공간 조금 더해서 설정 (2~3 정도 여유)
        if max_length > 0:
            ws.column_dimensions[col_letter].width = max_length + 2


# =====================================================================
# === STEP 2: PepMLM (ESM-2) 기반 펩타이드 생성 =======================
# =====================================================================

def load_esm_mlm(model_name: str = "facebook/esm2_t12_35M_UR50D"):
    print("\n" + "=" * 80)
    print("STEP 2: PepMLM (ESM-2) 모델 로딩")
    print("=" * 80)
    print(f"모델 로딩: {model_name}")
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = AutoModelForMaskedLM.from_pretrained(model_name).to(DEVICE)
    model.eval()
    print("✅ PepMLM 모델 로딩 완료")
    return tokenizer, model


def generate_peptides_with_mlm(
    tokenizer,
    model,
    target_sequence: str,
    num_peptides: int = NUM_PEPTIDES,
    peptide_len: int = PEPTIDE_LENGTH,
    top_k: int = 10,  # default 10
    temperature: float = 1.0,  # default 1.0
):
    """
    PepMLM(ESM-2 스타일) 기반 펩타이드 생성 (타깃 서열 컨텍스트 사용 버전)

    입력 프롬프트 형태:
      "[타깃 서열 토큰들] [MASK] [MASK] ... (펩타이드 길이만큼)"

    - 타깃 서열은 아미노산 한 글자씩 공백으로 나눈 형태로 토큰화:
        예) "AFK" -> "A F K"
    - 그 뒤에 [MASK] 토큰을 peptide_len 개 붙여서,
      "타깃 C-말단 뒤에 이어지는 펩타이드"를 모델이 채우도록 유도.

    top_k:
      - 각 MASK 위치에서 확률 상위 k개 아미노산만 남기고 샘플링.
      - k 작게 → 모델이 가장 그럴듯하다고 보는 아미노산 위주 (보수적).
      - k 크게 → 다양한 후보 (탐색적).

    temperature:
      - logits / temperature 후 softmax.
      - 1.0   → 원래 분포.
      - < 1.0 → 분포가 날카로워져서 고확률 토큰 위주 (덜 랜덤).
      - > 1.0 → 분포가 평탄해져서 저확률 토큰도 선택 (더 랜덤).

    이 버전은 이전의 "[PEP] [MASK]..." 프롬프트와 달리
    타깃 서열 전체를 컨텍스트로 사용하므로,
    "타깃 단백질과 어울리는 C-말단 확장 서열" 분포에 더 가깝게 생성한다.
    """

    print("\n펩타이드 서열 생성을 시작합니다 (타깃 컨텍스트 사용 버전)...")

    # 1) MASK 토큰 확인
    mask_token = tokenizer.mask_token
    if mask_token is None:
        raise ValueError("토크나이저에 [MASK] 토큰이 없습니다.")

    # 2) 제외할 토큰 ID (special token들은 샘플링 대상에서 제거)
    bad_ids = set()
    for tid in [
        tokenizer.pad_token_id,
        getattr(tokenizer, "cls_token_id", None),
        getattr(tokenizer, "sep_token_id", None),
        tokenizer.mask_token_id,
        getattr(tokenizer, "unk_token_id", None),
    ]:
        if tid is not None:
            bad_ids.add(tid)

    # 3) 타깃 서열을 아미노산 한 글자씩 공백으로 분리
    #    예: "AFKLV" → "A F K L V"
    target_seq = target_sequence.strip()
    target_tokens = " ".join(list(target_seq))

    # 4) 펩타이드 길이만큼 MASK 토큰을 뒤에 붙이기
    #    최종 프롬프트:
    #      "A F K L V [MASK] [MASK] ... [MASK]"
    mask_tokens = " ".join([mask_token] * peptide_len)
    prompt = f"{target_tokens} {mask_tokens}"

    input_ids = tokenizer(prompt, return_tensors="pt").input_ids.to(DEVICE)

    peptides = []
    seen = set()

    allowed_aas = set("ACDEFGHIKLMNPQRSTVWY")

    with torch.no_grad():
        attempt = 0
        # 최대 num_peptides 개 생성 시도 (중복/비표준 아미노산 때문에 여유를 둠)
        while len(peptides) < num_peptides and attempt < num_peptides * 5:
            attempt += 1
            ids = input_ids.clone()

            # 모든 [MASK] 위치를 순서대로 하나씩 채움
            for pos in range(ids.size(1)):
                if ids[0, pos].item() == tokenizer.mask_token_id:
                    outputs = model(ids)
                    logits = outputs.logits[0, pos] / temperature
                    probs = F.softmax(logits, dim=-1)

                    # special token 확률 0으로
                    for bid in bad_ids:
                        probs[bid] = 0.0

                    # 수치 오차 방지용 재정규화
                    probs_sum = probs.sum()
                    if probs_sum.item() == 0.0:
                        # 전부 0이 되면 이 위치 샘플링은 포기하고 다음 시도로
                        break
                    probs = probs / probs_sum

                    # top-k 필터링 후 샘플링
                    k = min(top_k, probs.size(0))
                    top_vals, top_idx = torch.topk(probs, k=k)
                    top_vals = top_vals / top_vals.sum()
                    sampled_local = torch.multinomial(top_vals, num_samples=1)
                    sampled_id = top_idx[sampled_local]
                    ids[0, pos] = sampled_id

            # 마스크를 모두 채운 뒤에 한 번만 디코딩
            seq = tokenizer.decode(ids[0], skip_special_tokens=True).replace(" ", "")

            # 프롬프트는 [타깃 서열 + 펩타이드] 구조이므로
            # 마지막 peptide_len 글자를 펩타이드로 잘라서 사용
            pep = seq[-peptide_len:]

            if len(pep) != peptide_len:
                continue

            # 표준 20개 아미노산만 허용 (X, B, Z, U, O, J 등 제외)
            if any(a not in allowed_aas for a in pep):
                continue

            # 중복 제거
            if pep in seen:
                continue

            seen.add(pep)
            peptides.append(pep)
            print(f"  [{len(peptides)}/{num_peptides}] 생성 완료: {pep} (길이: {len(pep)})")

    print("\n--- 생성된 펩타이드 후보 목록 ---")
    for i, p in enumerate(peptides, 1):
        print(f"  - 후보 {i}: {p}")
    print("=" * 80)
    print(f"STEP 2 완료: 총 {len(peptides)}개 펩타이드 후보 생성 (타깃 컨텍스트 사용)")
    print("=" * 80)
    return peptides


def write_target_fasta(fasta_dir: Path, target_sequence: str) -> Path:
    fasta_path = fasta_dir / "target_protein.fasta"
    with open(fasta_path, "w") as f:
        f.write(">target_protein\n")
        f.write(target_sequence.strip() + "\n")
    return fasta_path


def write_peptide_fasta(fasta_dir: Path, peptides) -> Path:
    pep_fa = fasta_dir / "peptides.fasta"
    with open(pep_fa, "w") as f:
        for i, pep in enumerate(peptides):
            f.write(f">pep_{i}\n{pep}\n")
    return pep_fa


# =====================================================================
# === STEP 3: ColabFold 배치 (멀티머) ================================
# =====================================================================

def prepare_colabfold_batch_csv(temp_dir: Path, target_sequence: str, peptides) -> Path:
    """
    ColabFold 1.5.5용 multimer CSV 입력 생성.

    - CSV 컬럼: id, sequence
    - sequence 형식: "타깃서열:펩타이드서열"
      예) AAAAA...AAAA:PPPP
    """
    csv_path = temp_dir / "batch_complexes.csv"
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "sequence"])
        for i, pep in enumerate(peptides):
            complex_id = f"complex_{i}"
            complex_seq = f"{target_sequence}:{pep}"
            writer.writerow([complex_id, complex_seq])
    print(f"✅ ColabFold 배치 CSV 생성 (id,sequence 형식): {csv_path}")
    return csv_path


def run_colabfold_batch_with_progress(
    csv_path: Path,
    out_dir: Path,
    total_complexes: int,
    max_msa: str = COLABFOLD_MAX_MSA,
    max_idle_min: int = COLABFOLD_MAX_IDLE_MIN,
    max_total_min: int = COLABFOLD_MAX_TOTAL_MIN,
):
    """
    colabfold_batch 실행 + 진행 상황 출력:
    - 기본은 GPU로 시도
    - GPU에서 RESOURCE_EXHAUSTED / Out of memory 발생 시,
      한 번에 한해 CPU(JAX_PLATFORM_NAME=cpu, CUDA_VISIBLE_DEVICES='')로 재시도

    진행 상황:
    - rank_001*.pdb 개수를 주기적으로 세어
      "완료된 구조 개수 / 전체 복합체 개수" 형태로 출력
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    cmd = [
        COLABFOLD_CMD,
        "--num-recycle", "1",
        "--model-type", "alphafold2_multimer_v3",
        "--rank", "ptm",
        "--max-msa", max_msa,
        "--num-models", "1",
        "--stop-at-score", "0.5",
        str(csv_path),
        str(out_dir),
    ]

    print("\n" + "=" * 80)
    print("STEP 3: ColabFold 배치 실행")
    print("=" * 80)
    print("[INFO] 실행 명령어:")
    print(" ", " ".join(cmd))

    # ColabFold 시작 전에 현재 프로세스에서 잡고 있는 GPU 메모리를 최대한 비워둔다.
    clear_gpu_memory()

    def _run_on_device(device_label: str, extra_env: dict | None, log_name: str):
        """
        실제로 colabfold_batch 를 한 번 실행하는 내부 함수.
        - device_label: 'GPU' 또는 'CPU' 등의 표시용 문자열
        - extra_env   : 특정 디바이스에서만 쓰고 싶은 환경변수 dict
        - log_name    : 로그 파일 이름
        """
        log_file = out_dir / log_name
        print(f"[INFO] ColabFold {device_label} 모드 실행")
        print(f"[INFO] 로그 파일: {log_file}")

        env = os.environ.copy()
        # 안전하게 XLA 메모리 설정 보장
        env.setdefault("XLA_PYTHON_CLIENT_PREALLOCATE", "false")
        env.setdefault(
            "XLA_PYTHON_CLIENT_MEM_FRACTION",
            os.environ.get("XLA_PYTHON_CLIENT_MEM_FRACTION", "0.8"),
        )
        if extra_env:
            env.update(extra_env)

        # colabfold_batch 실행
        with open(log_file, "w") as lf:
            proc = subprocess.Popen(cmd, stdout=lf, stderr=subprocess.STDOUT, env=env)

        start_time = time.time()
        last_done = -1
        last_progress_time = start_time

        # 진행 상황 모니터링 루프
        while True:
            ret = proc.poll()

            rank1_files = list(out_dir.glob("*rank_001*.*pdb"))
            done = len(rank1_files)
            if done != last_done:
                print(
                    f"\r[ColabFold 진행 상황({device_label})] {done}/{total_complexes} 구조 완료",
                    end="",
                    flush=True,
                )
                last_done = done
                last_progress_time = time.time()

            now = time.time()

            # 1) idle timeout: 진행률이 너무 오래 안 변하면 강제 종료
            if (now - last_progress_time) > max_idle_min * 60:
                print(
                    f"\n[ERROR] ColabFold({device_label})가 {max_idle_min}분 동안 "
                    "진행률이 변하지 않아 강제 중단합니다 (메모리 부족 또는 내부 오류 가능성)."
                )
                proc.terminate()
                try:
                    proc.wait(timeout=60)
                except subprocess.TimeoutExpired:
                    proc.kill()
                print(f"[INFO] 강제 종료 후 ColabFold 로그를 확인하세요: {log_file}")
                raise RuntimeError(
                    f"ColabFold({device_label}) 강제 종료 (idle timeout {max_idle_min}분 초과). "
                    f"로그: {log_file}"
                )

            # 2) 전체 실행 시간 상한 초과 시 강제 종료
            if (now - start_time) > max_total_min * 60:
                print(
                    f"\n[ERROR] ColabFold({device_label}) 전체 실행 시간이 "
                    f"{max_total_min}분을 넘어 강제 중단합니다."
                )
                proc.terminate()
                try:
                    proc.wait(timeout=60)
                except subprocess.TimeoutExpired:
                    proc.kill()
                print(f"[INFO] 강제 종료 후 ColabFold 로그를 확인하세요: {log_file}")
                raise RuntimeError(
                    f"ColabFold({device_label}) 강제 종료 (total timeout {max_total_min}분 초과). "
                    f"로그: {log_file}"
                )

            if ret is not None:
                break

            time.sleep(30)

        print()
        if proc.returncode != 0:
            print(
                f"[ERROR] ColabFold({device_label}) 실행 실패 "
                f"(returncode={proc.returncode}). 마지막 40줄 로그:"
            )
            try:
                with open(log_file) as f:
                    lines = f.readlines()
                tail_lines = lines[-40:]
                for line in tail_lines:
                    print(line.rstrip())
            except Exception as e:
                print(f"[WARN] 로그 파일을 읽는 중 오류 발생: {e}")
            raise RuntimeError(
                f"ColabFold({device_label}) 실행 실패, 로그 확인: {log_file}"
            )

        rank1_files = sorted(out_dir.glob("*rank_001*.*pdb"))
        print(f"[INFO] ColabFold({device_label}) 실행 완료")
        print(f"[INFO] rank_001 PDB 개수: {len(rank1_files)}")
        return rank1_files

    # 1차 시도: GPU 모드
    try:
        rank1_files = _run_on_device("GPU", extra_env=None, log_name="colabfold_batch.log")
    except RuntimeError:
        # GPU 모드에서 실패했을 때만 CPU fallback 고려
        if not COLABFOLD_CPU_FALLBACK:
            raise

        gpu_log_file = out_dir / "colabfold_batch.log"
        tail_text = ""
        try:
            with open(gpu_log_file) as f:
                lines = f.readlines()
            tail_text = "".join(lines[-80:])
        except Exception:
            pass

        # OOM 관련 키워드가 로그에 있는지 확인
        oom_keywords = (
            "RESOURCE_EXHAUSTED",
            "Out of memory",
            "out of memory",
            "CUDA_ERROR_OUT_OF_MEMORY",
        )
        is_oom = any(k in tail_text for k in oom_keywords)

        if is_oom:
            print("\n[WARN] GPU 메모리 부족(OOM)으로 ColabFold 실행 실패를 감지했습니다.")
            print("       CPU 모드(JAX_PLATFORMS=cpu, CUDA_VISIBLE_DEVICES='')로 한 번 더 재시도합니다.")
            cpu_env = {
                # GPU 완전 비활성화
                "CUDA_VISIBLE_DEVICES": "",
                # 새 JAX 버전용 (로그에서 직접 요구하던 설정)
                "JAX_PLATFORMS": "cpu",
                # 구버전 JAX 호환용 (혹시라도 쓰고 있을 수도 있어서 같이 넣어줌)
                "JAX_PLATFORM_NAME": "cpu",
            }
            # CPU 모드는 별도 로그 파일 이름 사용
            rank1_files = _run_on_device(
                "CPU",
                extra_env=cpu_env,
                log_name="colabfold_batch_cpu.log",
            )
        else:
            # OOM이 아닌 다른 이유라면 그대로 에러 전파
            raise


    # ColabFold 실행 후에도 혹시 남아 있을 수 있는 캐시 한 번 더 정리
    clear_gpu_memory()

    return rank1_files


# =====================================================================
# === STEP 4: AutoDock Vina 도킹 =====================================
# =====================================================================

# class ChainSelect(Select):
#     def __init__(self, chain_id):
#         self.chain_id = chain_id

#     def accept_chain(self, chain):
#         return chain.get_id() == self.chain_id


def split_complex_to_receptor_ligand(
    complex_pdb: Path,
    out_dir: Path,
    receptor_chain: str = "A",
    ligand_chain: str = "B",
):
    """
    ColabFold 멀티머 출력 PDB에서 지정한 체인 전체를 그대로 잘라
    receptor/ligand PDB를 만든다.
    - receptor_chain: 보통 A (타깃 단백질)
    - ligand_chain  : 보통 B (펩타이드 전체)
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    rec_pdb = out_dir / f"{complex_pdb.stem}_receptor_{receptor_chain}.pdb"
    lig_pdb = out_dir / f"{complex_pdb.stem}_ligand_{ligand_chain}.pdb"

    # 여기서 체인 전체를 그대로 사용
    extract_chain_as_pdb(complex_pdb, receptor_chain, rec_pdb)
    extract_chain_as_pdb(complex_pdb, ligand_chain, lig_pdb)

    return rec_pdb, lig_pdb


def compute_box_from_ligand(lig_pdb: Path, padding: float = 10.0):
    """
    리간드 좌표를 기반으로 박스 중심/크기를 자동 설정.
    """
    parser = PDBParser(QUIET=True)
    structure = parser.get_structure("ligand", str(lig_pdb))
    model = next(structure.get_models())
    coords = []
    for atom in model.get_atoms():
        coord = atom.get_coord()
        coords.append(coord)
    if not coords:
        raise ValueError(f"리간드 PDB에서 원자 좌표를 찾지 못했습니다: {lig_pdb}")

    coords = np.array(coords)
    center = coords.mean(axis=0)
    minc = coords.min(axis=0)
    maxc = coords.max(axis=0)
    size = (maxc - minc) + padding

    box = {
        "center_x": float(center[0]),
        "center_y": float(center[1]),
        "center_z": float(center[2]),
        "size_x": float(size[0]),
        "size_y": float(size[1]),
        "size_z": float(size[2]),
    }
    return box


def merge_ligand_roots(raw_pdbqt_path: str, out_pdbqt_path: str) -> None:
    """
    obabel이 만든 리간드 PDBQT(여러 ROOT/BRANCH 포함 가능)를
    Vina 1.2.x가 잘 읽을 수 있는 형태의
    단일 ROOT + TORSDOF 블록으로 변환한다.

    - REMARK 라인들은 그대로 유지
    - 모든 ATOM/HETATM 좌표를 하나의 ROOT/ENDROOT 안에 넣음
    - TORSDOF는 0(완전 rigid ligand)로 고정
    - !!! END 라인은 쓰지 않는다 (Vina 1.2.x에서 에러 남)
    """

    lines = Path(raw_pdbqt_path).read_text().splitlines()

    header_lines = []
    atom_lines = []

    # obabel이 계산한 torsdof 값은 참고만 하고 실제로는 0으로 rigid 처리
    orig_torsdof = None

    for line in lines:
        if not line.strip():
            continue

        tag = line[:6].strip()

        if tag == "REMARK":
            header_lines.append(line)
        elif tag in ("ATOM", "HETATM"):
            atom_lines.append(line)
        elif tag == "TORSDOF":
            # "TORSDOF 15" 이런 형태라서 뽑을 수는 있지만,
            # 여기서는 rigid ligand로 만들 거라 실제 값은 쓰지 않음.
            try:
                orig_torsdof = int(line.split()[1])
            except Exception:
                pass
        # BRANCH / ENDBRANCH / ROOT / ENDROOT / END 등은 전부 버림

    if not atom_lines:
        raise ValueError(f"No ATOM/HETATM lines found in ligand PDBQT: {raw_pdbqt_path}")

    out_lines = []
    out_lines.extend(header_lines)
    out_lines.append("ROOT")
    out_lines.extend(atom_lines)
    out_lines.append("ENDROOT")
    # 리간드 rigid 처리
    out_lines.append("TORSDOF 0")

    # 여기에서 절대 "END" 를 추가하지 않는다!
    Path(out_pdbqt_path).write_text("\n".join(out_lines) + "\n")


def sanitize_receptor_pdbqt(rec_pdbqt: Path):
    """
    리셉터 PDBQT에서 Vina가 싫어할 만한 태그들을 제거하고
    REMARK / ATOM / HETATM / END 정도만 남긴다.

    - ROOT, ENDROOT, BRANCH, ENDBRANCH, TORSDOF, CONECT, MODEL, ENDMDL 등은 모두 제거.
    """
    if not rec_pdbqt.exists():
        print(f"[WARN] sanitize_receptor_pdbqt: 파일 없음: {rec_pdbqt}")
        return

    lines_in = rec_pdbqt.read_text().splitlines()
    out_lines = []

    for line in lines_in:
        s = line.strip()
        if not s:
            # 빈 줄은 그냥 버리자
            continue

        # 리셉터에선 필요 없는 태그들 싹 제거
        if s.startswith(("ROOT", "ENDROOT", "BRANCH", "ENDBRANCH", "TORSDOF",
                         "CONECT", "MODEL", "ENDMDL")):
            continue

        # REMARK / ATOM / HETATM / END 만 허용
        if s.startswith(("REMARK", "ATOM", "HETATM", "END")):
            out_lines.append(s)
            continue

        # 그 외 태그는 전부 버림
        continue

    rec_pdbqt.write_text("\n".join(out_lines) + "\n")


def prepare_pdbqt(rec_pdb: Path, lig_pdb: Path, out_dir: Path) -> tuple[Path, Path]:
    """
    - receptor: obabel -xr 로 PDBQT 생성 후, 혹시 모를 ROOT/BRANCH 태그는 제거
    - ligand  : obabel로 임시 PDBQT 생성 → merge_ligand_roots 로 단일 ROOT rigid ligand 구성
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    rec_pdbqt = out_dir / f"{rec_pdb.stem}.pdbqt"
    lig_raw_pdbqt = out_dir / f"{lig_pdb.stem}_raw.pdbqt"
    lig_pdbqt = out_dir / f"{lig_pdb.stem}.pdbqt"

    # 1) receptor PDBQT (rigid)
    rec_cmd = [
        OBABEL_CMD,
        "-ipdb", str(rec_pdb),
        "-xr",                    # receptor 모드
        "-opdbqt",
        "-O", str(rec_pdbqt),
    ]
    print("[RUN]", " ".join(rec_cmd))
    rec_res = subprocess.run(rec_cmd, capture_output=True, text=True)
    if rec_res.returncode != 0:
        print(rec_res.stdout)
        print(rec_res.stderr)
        raise RuntimeError(f"receptor PDBQT 변환 실패: code={rec_res.returncode}")

    # 안전하게 ROOT/BRANCH 등 제거
    sanitize_receptor_pdbqt(rec_pdbqt)

    # 2) ligand PDBQT (obabel → 단일 ROOT rigid ligand)
    lig_cmd = [
        OBABEL_CMD,
        "-ipdb", str(lig_pdb),
        "-opdbqt",
        "-O", str(lig_raw_pdbqt),
    ]
    print("[RUN]", " ".join(lig_cmd))
    lig_res = subprocess.run(lig_cmd, capture_output=True, text=True)
    if lig_res.returncode != 0:
        print(lig_res.stdout)
        print(lig_res.stderr)
        raise RuntimeError(f"ligand PDBQT 변환 실패: code={lig_res.returncode}")

    # 여러 ROOT/모델을 하나의 rigid ligand로 합치기
    merge_ligand_roots(lig_raw_pdbqt, lig_pdbqt)

    # 원본 임시 파일은 필요 없으면 삭제해도 됨
    try:
        lig_raw_pdbqt.unlink()
    except OSError:
        pass

    return rec_pdbqt, lig_pdbqt


def parse_vina_score_from_stdout(stdout: str):
    """
    AutoDock Vina stdout에서 best score(affinity, kcal/mol)를 파싱.

    우선순위:
    1) mode 테이블 (mode | affinity | ...)에서 affinity 열 파싱
    2) fallback: 'REMARK VINA RESULT' 형식이 있으면 그 줄에서 float 추출
    """
    energies = []

    # 1) mode 테이블 파싱
    for line in stdout.splitlines():
        s = line.strip()
        if not s:
            continue
        # 헤더/구분선은 건너뛰기
        if s.startswith("mode") or set(s) <= {"-", "+"}:
            continue

        parts = s.split()
        # "1  -7.5  ..." 이런 형식일 때
        if parts and parts[0].isdigit() and len(parts) >= 2:
            try:
                val = float(parts[1])
            except ValueError:
                continue
            energies.append(val)

    if energies:
        # 가장 낮은 에너지(가장 좋은 포즈)를 반환
        return min(energies)

    # 2) fallback: 예전 스타일 'REMARK VINA RESULT:' 줄
    for line in stdout.splitlines():
        if "REMARK VINA RESULT" in line:
            for token in line.split():
                try:
                    return float(token)
                except ValueError:
                    continue

    # 아무 것도 못 찾으면 None
    return None


def get_chain_residue_counts(pdb_path: Path):
    """
    PDB 파일에서 체인별 residue 개수를 세어 반환.
    - key: 체인 ID (문자)
    - value: 해당 체인의 고유 residue 개수
    """
    chain_res = defaultdict(set)
    with open(pdb_path, "r") as f:
        for line in f:
            if not (line.startswith("ATOM") or line.startswith("HETATM")):
                continue
            if len(line) < 26:
                continue
            chain_id = line[21].strip() or "_"
            res_seq = line[22:26].strip()
            if not res_seq:
                continue
            chain_res[chain_id].add(res_seq)
    return {cid: len(res) for cid, res in chain_res.items()}


def auto_assign_receptor_ligand(chain_counts: dict, prefer_receptor: str = "A"):
    """
    체인별 residue 개수를 바탕으로 receptor/ligand 체인을 추정.
    - prefer_receptor 가 chain_counts에 있으면 → 그 체인을 receptor로 확정
      ligand는 나머지 체인 중 residue 개수가 가장 적은 체인
    - prefer_receptor가 없으면 → residue 개수가 가장 많은 체인을 receptor,
      가장 적은 체인을 ligand로 선택
    - 체인이 1개뿐이면 (receptor, None) 반환
    """
    chains = list(chain_counts.keys())
    if not chains:
        return None, None
    if len(chains) == 1:
        # 단일체 구조 → ligand 없음
        return chains[0], None

    if prefer_receptor in chain_counts:
        receptor = prefer_receptor
        others = [c for c in chains if c != receptor]
        ligand = min(others, key=lambda c: chain_counts[c])
    else:
        # 가장 큰 체인을 receptor, 가장 작은 체인을 ligand 로
        receptor = max(chains, key=lambda c: chain_counts[c])
        others = [c for c in chains if c != receptor]
        ligand = min(others, key=lambda c: chain_counts[c])

    return receptor, ligand


def run_vina_on_rank1(rank1_pdbs, vina_dir: Path):
    """
    AutoDock Vina 도킹 실행 + 요약/상태 로그 기록.

    - vina_summary.xlsx 컬럼:
        complex, vina_score, vina_status, receptor_pdbqt, ligand_pdbqt, log_file
    - vina_status 예시:
        '정상', '정상(점수=0.0)',
        '실패: Vina 실행 에러(code=...)',
        '실패: Vina 실행 에러(code=...)(PDBQT parsing error: ...)',
        '파싱실패: stdout에서 점수 패턴 없음',
        '스킵: 단일체 구조(리간드 체인 없음)',
        '스킵: 리간드 체인 자동 탐지 실패',
        '스킵: 체인 분리 실패(...)'
    """
    print("\n" + "=" * 80)
    print("STEP 4: AutoDock Vina 도킹")
    print("=" * 80)

    if not rank1_pdbs:
        print("[WARN] Vina 실행할 rank_001 PDB가 없습니다.")
        return
    if not shutil.which(VINA_CMD):
        print(f"[WARN] VINA_CMD='{VINA_CMD}' 실행 파일을 찾을 수 없습니다. (PATH 확인 필요)")
        return

    vina_dir.mkdir(parents=True, exist_ok=True)
    summary_rows = []  # 딕셔너리 리스트로 관리
    debug_lines = []

    for complex_pdb in rank1_pdbs:
        base = complex_pdb.stem
        complex_out_dir = vina_dir / base
        complex_out_dir.mkdir(parents=True, exist_ok=True)

        log_file = complex_out_dir / f"{base}_vina_stdout.txt"

        # 이 한 줄로 log_file까지 포함해 기본 row 구조를 잡아놓고 시작
        row_data = {
            "complex": base,
            "vina_status": "",
            "vina_score": None,
            "receptor_pdbqt": "",
            "ligand_pdbqt": "",
            "log_file": log_file.name,
        }

        print(f"\n[INFO] Vina 준비: {complex_pdb.name}")

        # 1) PDB 체인 구성 분석
        chain_counts = get_chain_residue_counts(complex_pdb)
        print(f"[INFO] {complex_pdb.name} 체인 구성: {chain_counts}")

        if not chain_counts:
            status = "스킵: 체인 정보 없음(ATOM/HETATM 레코드 없음)"
            print(f"[WARN] {status}")
            log_file.write_text(status + "\n", encoding="utf-8")
            row_data["vina_status"] = status
            summary_rows.append(row_data)
            debug_lines.append(f"{base}\t{status}\tlog={log_file.name}")
            continue

        # 체인 수가 1개면 Vina 도킹 불가 → 스킵
        if len(chain_counts) == 1:
            status = "스킵: 단일체 구조(리간드 체인 없음)"
            print(f"[WARN] {complex_pdb.name} {status}")
            msg = f"{status}\nchains={chain_counts}\n"
            log_file.write_text(msg, encoding="utf-8")
            row_data["vina_status"] = status
            summary_rows.append(row_data)
            debug_lines.append(f"{base}\t{status}\tlog={log_file.name}")
            continue

        # 2) receptor / ligand 체인 자동 결정
        rec_chain, lig_chain = auto_assign_receptor_ligand(chain_counts, prefer_receptor="A")
        if rec_chain is None or lig_chain is None:
            status = "스킵: 리간드 체인 자동 탐지 실패"
            print(f"[WARN] {complex_pdb.name} {status}")
            msg = f"{status}\nchains={chain_counts}\n"
            log_file.write_text(msg, encoding="utf-8")
            row_data["vina_status"] = status
            summary_rows.append(row_data)
            debug_lines.append(f"{base}\t{status}\tlog={log_file.name}")
            continue

        print(f"[INFO] 자동 할당 체인: receptor={rec_chain}, ligand={lig_chain}")

        # 3) 체인 분리 (예전에는 A/B로 고정이어서 여기서 ValueError가 터졌던 부분)
        try:
            rec_pdb, lig_pdb = split_complex_to_receptor_ligand(
                complex_pdb,
                complex_out_dir,
                receptor_chain=rec_chain,
                ligand_chain=lig_chain,
            )
        except ValueError as e:
            status = f"스킵: 체인 분리 실패({e})"
            print(f"[WARN] {complex_pdb.name} {status}")
            log_msg = f"{status}\nchains={chain_counts}\n"
            log_file.write_text(log_msg, encoding="utf-8")
            row_data["vina_status"] = status
            summary_rows.append(row_data)
            debug_lines.append(f"{base}\t{status}\tlog={log_file.name}")
            continue

        # 4) PDBQT 준비
        rec_pdbqt, lig_pdbqt = prepare_pdbqt(rec_pdb, lig_pdb, complex_out_dir)
        row_data["receptor_pdbqt"] = rec_pdbqt.name
        row_data["ligand_pdbqt"] = lig_pdbqt.name
        box = compute_box_from_ligand(lig_pdb)

        out_pdbqt = complex_out_dir / f"{base}_vina_out.pdbqt"

        vina_cmd = (
            f"{VINA_CMD} "
            f"--receptor {rec_pdbqt} "
            f"--ligand {lig_pdbqt} "
            f"--center_x {box['center_x']:.3f} --center_y {box['center_y']:.3f} --center_z {box['center_z']:.3f} "
            f"--size_x {box['size_x']:.3f} --size_y {box['size_y']:.3f} --size_z {box['size_z']:.3f} "
            f"--out {out_pdbqt}"
        )

        print(f"[RUN] {vina_cmd}")
        result = subprocess.run(
            vina_cmd,
            shell=True,
            capture_output=True,
            text=True,
        )

        with open(log_file, "w", encoding="utf-8") as lf:
            lf.write("=== STDOUT ===\n")
            lf.write(result.stdout or "")
            lf.write("\n\n=== STDERR ===\n")
            lf.write(result.stderr or "")

        best_score = None
        status = ""

        if result.returncode != 0:
            # 실행 자체 실패
            status = f"실패: Vina 실행 에러(code={result.returncode})"
            if "PDBQT parsing error" in (result.stdout or "") or "PDBQT parsing error" in (result.stderr or ""):
                status += " (PDBQT parsing error: flex residue/ligand 태그 문제)"

                # 디버그: ligand PDBQT 안에 어떤 태그들이 있는지 확인
                try:
                    tags = set()
                    with open(lig_pdbqt, "r") as lf:
                        for ln in lf:
                            t = ln.strip()
                            if not t:
                                continue
                            first = t.split()[0]
                            tags.add(first)
                    print(f"[DEBUG] {base} ligand_pdbqt 태그 목록:", sorted(tags))
                except Exception as e:
                    print(f"[DEBUG] {base} ligand_pdbqt 태그 읽기 실패: {e}")

            print(f"[ERROR] {status}. 로그 파일: {log_file}")
            print(result.stdout or "")
            print(result.stderr or "")
            debug_lines.append(f"{base}\t{status}\tlog={log_file.name}")
        else:
            # 실행은 성공 → score 파싱
            best_score = parse_vina_score_from_stdout(result.stdout or "")
            if best_score is None:
                status = "파싱실패: stdout에서 점수 패턴 없음"
                print(f"[WARN] {complex_pdb.name} Vina 점수 파싱 실패. 로그 파일 확인: {log_file}")
                debug_lines.append(f"{base}\t{status}\tlog={log_file.name}")
            else:
                status = "정상"
                print(f"[INFO] {complex_pdb.name} Vina best score: {best_score}")
                debug_lines.append(f"{base}\t{status}\t{best_score}")

        row_data["vina_status"] = status
        row_data["vina_score"] = best_score
        summary_rows.append(row_data)

    # 요약 엑셀 저장 (첫 행은 헤더)
    try:
        if summary_rows:
            df_vina = pd.DataFrame(summary_rows)
            # 컬럼 순서를 상수로 강제
            df_vina = df_vina[VINA_SUMMARY_COLS]
            xlsx_path = vina_dir / "vina_summary.xlsx"
            df_vina.to_excel(xlsx_path, index=False)
            print(f"\n✅ Vina 요약 엑셀 저장: {xlsx_path}")
        else:
            print("[INFO] Vina 요약에 기록할 데이터가 없습니다.")
    except Exception as e:
        print(f"[WARN] Vina 요약 엑셀 저장 실패: {e}")

    if debug_lines:
        debug_file = vina_dir / "vina_debug.txt"
        debug_file.write_text("\n".join(debug_lines), encoding="utf-8")
        print(f"[INFO] Vina 디버그 로그: {debug_file}")

    print("=" * 80)


# =====================================================================
# === STEP 5: PLIP 상호작용 분석 =====================================
# =====================================================================

def run_plip_on_rank1(rank1_pdbs, plip_dir: Path):
    """
    PLIP 상호작용 분석.
    - 각 complex별 stdout/stderr 를 로그 파일로 저장.
    """
    print("\n" + "=" * 80)
    print("STEP 5: PLIP 상호작용 분석")
    print("=" * 80)
    if not rank1_pdbs:
        print("[WARN] PLIP 실행할 rank_001 PDB가 없습니다.")
        return
    if not PLIP_CMD:
        print("[WARN] PLIP_CMD 가 비어 있습니다.")
        return

    plip_dir.mkdir(parents=True, exist_ok=True)
    debug_lines = []

    for pdb in rank1_pdbs:
        base = pdb.stem
        out_subdir = plip_dir / base
        out_subdir.mkdir(exist_ok=True)

        log_file = out_subdir / f"{base}_plip.log"

        # 체인 구성 파악
        chain_counts = get_chain_residue_counts(pdb)
        rec_chain, lig_chain = auto_assign_receptor_ligand(chain_counts, prefer_receptor="A")

        if rec_chain is None or lig_chain is None:
            status = "스킵: PLIP용 리간드 체인 자동 탐지 실패"
            msg = f"{status}\nchains={chain_counts}\n"
            log_file.write_text(msg, encoding="utf-8")
            print(f"[WARN] {pdb.name} {status}")
            debug_lines.append(f"{base}\t{status}\tlog={log_file.name}")
            continue

        chains_expr = f"[['{rec_chain}'], ['{lig_chain}']]"

        cmd_list = [
            *PLIP_CMD.split(),
            "-f", str(pdb),
            "-o", str(out_subdir),
            "-x", "-t",
            "--chains", chains_expr,
        ]
        print("[RUN]", " ".join(cmd_list))

        result = subprocess.run(cmd_list, capture_output=True, text=True)

        with open(log_file, "w", encoding="utf-8") as lf:
            lf.write("=== STDOUT ===\n")
            lf.write(result.stdout or "")
            lf.write("\n\n=== STDERR ===\n")
            lf.write(result.stderr or "")

        if result.returncode != 0:
            print(f"[ERROR] PLIP 실패: {pdb.name} (code={result.returncode}). 로그: {log_file}")
            print((result.stderr or "")[:300])
            debug_lines.append(f"{base}\tERROR_returncode_{result.returncode}\tlog={log_file.name}")
        else:
            print(f"✔️ PLIP 완료: {pdb.name} → {out_subdir}")
            debug_lines.append(f"{base}\tOK\tlog={log_file.name}")

    if debug_lines:
        debug_file = plip_dir / "plip_run_debug.txt"
        debug_file.write_text("\n".join(debug_lines), encoding="utf-8")
        print(f"\n[INFO] PLIP 실행 디버그 로그: {debug_file}")

    # PLIP 결과 요약 파일(plip_summary.xlsx) 생성
    try:
        load_plip_scores(plip_dir)
    except Exception as e:
        print(f"[WARN] PLIP 요약 생성(load_plip_scores) 중 오류: {e}")

    print("=" * 80)


# =====================================================================
# === STEP 6: PRODIGY 결합 친화도 평가 ===============================
# =====================================================================

def run_prodigy_on_rank1(rank1_pdbs, out_dir: Path) -> pd.DataFrame:
    """
    PRODIGY 결합 친화도 평가.

    - prodigy_summary.xlsx 컬럼:
        complex, PRODIGY_status, PRODIGY_dG
    """
    print("\n" + "="*80)
    print("STEP 6: PRODIGY 결합 친화도 평가")
    print("="*80)

    if not PRODIGY_SCRIPT:
        print("[WARN] PRODIGY_SCRIPT 환경변수가 설정되어 있지 않습니다.")
        print("       예: export PRODIGY_SCRIPT='prodigy'")
        return pd.DataFrame()

    out_dir.mkdir(parents=True, exist_ok=True)

    records = []
    debug_lines = []

    for pdb_path in rank1_pdbs:
        complex_name = Path(pdb_path).stem
        out_txt = out_dir / f"{complex_name}_prodigy.txt"
        err_txt = out_dir / f"{complex_name}_prodigy.stderr.txt"

        # 체인 구성 파악
        chain_counts = get_chain_residue_counts(pdb_path)
        rec_chain, lig_chain = auto_assign_receptor_ligand(chain_counts, prefer_receptor="A")

        if rec_chain is None or lig_chain is None:
            status = "스킵: PRODIGY용 리간드 체인 자동 탐지 실패"
            records.append({
                "complex": complex_name,
                "PRODIGY_status": status,
                "PRODIGY_dG": None,
            })
            debug_lines.append(f"{complex_name}\t{status}\tNone")
            msg = f"{status}\nchains={chain_counts}\n"
            out_txt.write_text(msg, encoding="utf-8")
            err_txt.write_text("", encoding="utf-8")
            continue

        cmd = [
            *PRODIGY_SCRIPT.split(),
            str(pdb_path),
            "--selection", rec_chain, lig_chain,
        ]
        print(f"[RUN] {' '.join(cmd)}")
        result = subprocess.run(cmd, capture_output=True, text=True)

        # stdout / stderr 저장
        out_txt.write_text(result.stdout or "", encoding="utf-8")
        err_txt.write_text(result.stderr or "", encoding="utf-8")

        dg = None
        status = ""

        if result.returncode != 0:
            status = f"실패: PRODIGY 실행 에러(code={result.returncode})"
            print(f"[ERROR] {status}: {complex_name}. 로그: {out_txt.name}, {err_txt.name}")
            print((result.stderr or "")[:300])
        else:
            stdout_text = result.stdout or ""
            lower = stdout_text.lower()

            # 1) 선택한 체인 사이에 접촉이 전혀 없는 경우
            if "no contacts found for selection" in lower:
                status = "실패: No contacts found for selection(선택된 체인 간 접촉 없음)"
                dg = None
                print(f"[WARN] PRODIGY 접촉 없음: {complex_name}. (로그: {out_txt.name})")

            # 2) 접촉은 있는데 ΔG 라인이 없어서 파싱 실패한 경우
            else:
                dg = parse_prodigy_dg_from_stdout(stdout_text)
                if dg is None:
                    status = "파싱실패: stdout에서 ΔG 패턴 없음"
                    print(f"[WARN] PRODIGY ΔG 파싱 실패: {complex_name}. (로그: {out_txt.name})")
                else:
                    status = "정상"

        records.append({
            "complex": complex_name,
            "PRODIGY_status": status,
            "PRODIGY_dG": dg,
        })
        debug_lines.append(f"{complex_name}\t{status}\t{dg}")

    if debug_lines:
        debug_file = out_dir / "prodigy_debug.txt"
        debug_file.write_text("\n".join(debug_lines), encoding="utf-8")
        print(f"[INFO] PRODIGY 디버그 로그: {debug_file}")

    df = pd.DataFrame(records)

    xlsx_path = out_dir / "prodigy_summary.xlsx"
    try:
        if not df.empty:
            df = df[PRODIGY_SUMMARY_COLS]
            df.to_excel(xlsx_path, index=False)
            print(f"✅ PRODIGY 요약 엑셀 저장: {xlsx_path}")
        else:
            print("[INFO] PRODIGY 요약에 기록할 데이터가 없습니다.")
    except Exception as e:
        print(f"[WARN] PRODIGY 요약 엑셀 저장 실패: {e}")

    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active  # prodigy_summary는 시트 하나뿐

        # PRODDIGY_dG 컬럼이 3번째라는 가정 (complex, PRODIGY_status, PRODIGY_dG)
        for cell in ws["C"]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

        wb.save(xlsx_path)
    except Exception as e:
        print(f"[WARN] PRODIGY 요약 엑셀 number_format 설정 실패: {e}")


    return df


# =====================================================================
# === STEP 7: 최종 평가(가중치 A안) + PDB zip + 엑셀 =================
# =====================================================================

def zip_rank1_pdbs(rank1_pdbs, results_dir: Path):
    """
    ColabFold에서 생성된 rank_001 PDB들을 하나의 zip 파일로 압축.
    (타깃 단백질 + 생성 펩타이드 복합체 구조)
    """
    if not rank1_pdbs:
        print("[INFO] zip으로 묶을 rank_001 PDB가 없습니다.")
        return None

    zip_path = results_dir / f"peptide_structures_{timestamp()}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for pdb_path in rank1_pdbs:
            zf.write(pdb_path, arcname=pdb_path.name)

    print(f"✅ rank_001 PDB 압축 저장: {zip_path}")
    return zip_path


def load_vina_scores(vina_dir: Path):
    """
    vina_summary.xlsx 에서 complex별 Vina score와 상태를 로딩.

    반환:
      scores   : dict[complex_id] = float 또는 None
      statuses : dict[complex_id] = 상태 문자열
    """
    scores = {}
    statuses = {}

    summary_xlsx = vina_dir / "vina_summary.xlsx"
    df = None

    if summary_xlsx.exists():
        try:
            df = pd.read_excel(summary_xlsx)
            print(f"[INFO] Vina 요약 엑셀에서 점수 로드: {summary_xlsx}")
        except Exception as e:
            print(f"[WARN] Vina 엑셀 로딩 실패: {e}")
    else:
        print("[WARN] Vina summary 파일(vina_summary.xlsx)를 찾지 못했습니다:", vina_dir)
        return scores, statuses

    if df is None or "complex" not in df.columns:
        print("[WARN] Vina summary 데이터프레임에 'complex' 컬럼이 없습니다.")
        return scores, statuses

    has_status = "vina_status" in df.columns

    for _, row in df.iterrows():
        base = row.get("complex")
        if isinstance(base, float) and pd.isna(base):
            continue
        base = str(base).strip()
        if not base:
            continue

        val = row.get("vina_score")

        # NaN 은 None 으로 간주
        if pd.isna(val):
            scores[base] = None
        else:
            try:
                scores[base] = float(val)
            except (TypeError, ValueError):
                scores[base] = None

        if has_status:
            s = row.get("vina_status")
            statuses[base] = "" if pd.isna(s) else str(s)
        else:
            statuses[base] = "정상" if scores[base] is not None else "미기록"

    print(f"[INFO] Vina 점수를 읽어온 구조 수: {len(scores)}")
    return scores, statuses


def load_prodigy_scores(prodigy_dir: Path):
    """
    PRODIGY 결과 로딩.

    우선순위:
      1) prodigy_summary.xlsx (있다면)
      2) 없으면 *_prodigy.txt 백업 파싱
    """
    scores = {}
    statuses = {}

    if not prodigy_dir.exists():
        return scores, statuses

    df = None

    summary_xlsx = prodigy_dir / "prodigy_summary.xlsx"
    try:
        df = pd.read_excel(summary_xlsx)
        print(f"[INFO] PRODIGY 요약 엑셀에서 점수 로드: {summary_xlsx}")
    except Exception as e:
        print(f"[WARN] prodigy_summary.xlsx 로딩 실패: {e}")

    if df is not None and "complex" in df.columns:
        # 값 컬럼(prodgy_dg, PRODIGY_dG 등) / 상태 컬럼(prodigy_status, PRODIGY_status 등) 자동 탐색
        val_col = None
        status_col = None
        for c in df.columns:
            cl = c.lower()
            if cl.startswith("prodigy"):
                if "status" in cl and status_col is None:
                    status_col = c
                elif "status" not in cl and val_col is None:
                    val_col = c

        for _, row in df.iterrows():
            comp = row.get("complex")
            if isinstance(comp, float) and pd.isna(comp):
                continue
            comp = str(comp).strip()
            if not comp:
                continue

            val = row.get(val_col) if val_col is not None else None

            # NaN → None
            if val is None or pd.isna(val):
                scores[comp] = None
            else:
                try:
                    scores[comp] = float(val)
                except (TypeError, ValueError):
                    scores[comp] = None

            if status_col is not None:
                s = row.get(status_col)
                statuses[comp] = "" if pd.isna(s) else str(s)
            else:
                statuses[comp] = "정상" if scores[comp] is not None else "미기록"

        print(f"[INFO] PRODIGY ΔG를 요약 파일에서 불러옴: {len(scores)}개 구조")

    # summary에서 뭔가 읽어왔다면 그대로 사용
    if scores or statuses:
        return scores, statuses

    # 백업: *_prodigy.txt 직접 파싱
    for txt in prodigy_dir.glob("*_prodigy.txt"):
        base = txt.stem.replace("_prodigy", "")
        try:
            text = txt.read_text()
        except Exception:
            continue

        vals = []
        for m in re.finditer(r"[-+]?\d+\.\d+", text):
            v = float(m.group(0))
            # PRODIGY ΔG 대략적인 범위
            if -50.0 <= v <= 0.0:
                vals.append(v)

        if vals:
            scores[base] = min(vals)
            statuses[base] = "텍스트 파싱(백업)"
        else:
            scores[base] = None
            statuses[base] = "파싱실패: *_prodigy.txt에서 ΔG 후보값 없음"

    print(f"[INFO] PRODIGY ΔG를 텍스트에서 직접 파싱: {len(scores)}개 구조")
    return scores, statuses


def load_iptm_scores(colabfold_out_dir: Path, rank1_pdbs):
    """
    ColabFold 출력 폴더에서 ipTM 값을 최대한 유연하게 찾는다.

    - 각 rank_001 PDB의 stem(base)를 기준으로
      1) base*scores*.json
      2) base_prefix*scores*.json  (base에서 '_unrelaxed' 앞부분)
      3) base*_ranking_debug.json, base_prefix*_ranking_debug.json, ranking_debug.json
    에서 'iptm' 또는 'iptm+ptm' 키를 찾아본다.
    """
    iptms = {}
    if not colabfold_out_dir.exists():
        return iptms

    for pdb in rank1_pdbs:
        base = pdb.stem
        prefix = base.split("_unrelaxed")[0]

        found_val = None

        # 1) scores*.json 후보들
        candidates = list(colabfold_out_dir.glob(f"{base}*scores*.json"))
        if not candidates:
            candidates = list(colabfold_out_dir.glob(f"{prefix}*scores*.json"))

        for js in candidates:
            try:
                with open(js) as f:
                    data = json.load(f)
            except Exception:
                continue

            if isinstance(data, dict):
                v = data.get("iptm")
                if isinstance(v, (int, float)):
                    found_val = float(v)
                    break
                v = data.get("iptm+ptm")
                if isinstance(v, (int, float)):
                    found_val = float(v)
                    break

        # 2) ranking_debug 후보들
        if found_val is None:
            rd_candidates = [
                colabfold_out_dir / f"{base}_ranking_debug.json",
                colabfold_out_dir / f"{prefix}_ranking_debug.json",
                colabfold_out_dir / "ranking_debug.json",
            ]
            for rd in rd_candidates:
                if not rd.exists():
                    continue
                try:
                    with open(rd) as f:
                        data = json.load(f)
                except Exception:
                    continue

                if isinstance(data, dict):
                    v = data.get("iptm") or data.get("iptm+ptm")
                    if isinstance(v, (int, float)):
                        found_val = float(v)
                        break

        if found_val is not None:
            iptms[base] = found_val

    print(f"[INFO] ipTM 값을 읽어온 구조 수: {len(iptms)} / {len(rank1_pdbs)}")
    return iptms


def load_plip_scores(plip_dir: Path):
    """
    PLIP 결과 폴더들에서 상호작용 스코어 + 상태를 추출.

    - 각 complex별 서브폴더 이름이 complex_0_unrelaxed_... 형태라고 가정.
    - 폴더 안/하위폴더 전체에서 *.xml / *.txt 파일을 찾아
      PLIP가 생성한 report 파일로부터 상호작용 수를 추출한다.

    반환:
      metrics  : dict[base] = {
                    "total": total or None,
                    "hbond": ...,
                    "hydrophobic": ...,
                    "saltbridge": ...
                }
      statuses : dict[base] = 상태 문자열
    """
    metrics = {}
    statuses = {}

    if not plip_dir.exists():
        return metrics, statuses

    debug_lines = []

    for subdir in plip_dir.iterdir():
        if not subdir.is_dir():
            continue

        base = subdir.name

        hbond = 0
        hydrophobic = 0
        saltbridge = 0
        total = None

        source = None
        status = ""

        # 1) report xml 후보들 (재귀적으로 *.xml 전체 탐색)
        xml_candidates = sorted(subdir.rglob("*.xml"))
        # 파일명에 'report'/'plip' 가 포함된 것을 우선
        xml_candidates.sort(
            key=lambda p: (
                "report" not in p.name.lower() and "plip" not in p.name.lower(),
                p.name,
            )
        )

        for xml_path in xml_candidates:
            try:
                tree = ET.parse(xml_path)
                root = tree.getroot()
                hbond = sum(1 for _ in root.iter("hydrogen_bond"))
                hydrophobic = sum(1 for _ in root.iter("hydrophobic_interaction"))
                saltbridge = sum(1 for _ in root.iter("salt_bridge"))
                total = hbond + hydrophobic + saltbridge
                source = f"xml({xml_path.name})"
                status = "정상"
                break
            except Exception as e:
                debug_lines.append(
                    f"{base}\treport_xml_parse_error({xml_path.name})\t{e}"
                )

        # 2) 텍스트 report/log 후보들 (재귀적으로 *.txt, *.log 전체 탐색)
        if source is None:
            txt_candidates = sorted(
                list(subdir.rglob("*.txt")) + list(subdir.rglob("*.log"))
            )
            # 파일명에 report 가 들어간 파일을 우선
            txt_candidates.sort(
                key=lambda p: (
                    "report" not in p.name.lower(),
                    p.name,
                )
            )

            for txt_report in txt_candidates:
                try:
                    text = txt_report.read_text()
                    hb = 0
                    hp = 0
                    sb = 0
                    found_any = False
                    for line in text.splitlines():
                        lower = line.lower()
                        nums = re.findall(r"\b\d+\b", line)
                        if not nums:
                            continue
                        last_num = int(nums[-1])
                        if "hydrogen bond" in lower:
                            hb = last_num
                            found_any = True
                        elif "hydrophobic" in lower:
                            hp = last_num
                            found_any = True
                        elif "salt bridge" in lower:
                            sb = last_num
                            found_any = True
                    # 상호작용 라인 하나라도 찾으면 성공으로 간주 (0개여도 가능)
                    if found_any:
                        hbond = hb
                        hydrophobic = hp
                        saltbridge = sb
                        total = hbond + hydrophobic + saltbridge
                        source = f"txt({txt_report.name})"
                        status = "정상(txt/log)"
                        break
                except Exception as e:
                    debug_lines.append(
                        f"{base}\treport_txt_parse_error({txt_report.name})\t{e}"
                    )

        # 3) 둘 다 실패하면 상태 메시지
        if source is None and not status:
            status = "실패: PLIP 요약 파일(xml/txt)을 찾지 못함"

        # total이 없거나 실패 상태면 세부값을 None으로
        if total is None or status.startswith("실패"):
            metrics[base] = {
                "total": None,
                "hbond": None,
                "hydrophobic": None,
                "saltbridge": None,
            }
        else:
            metrics[base] = {
                "total": total,
                "hbond": hbond,
                "hydrophobic": hydrophobic,
                "saltbridge": saltbridge,
            }

        statuses[base] = status
        debug_lines.append(
            f"{base}\t{status}\t"
            f"total={metrics[base]['total']}\t"
            f"hbond={metrics[base]['hbond']}\t"
            f"hydrophobic={metrics[base]['hydrophobic']}\t"
            f"saltbridge={metrics[base]['saltbridge']}"
        )

    # 디버그 로그
    if debug_lines:
        debug_file = plip_dir / "plip_parse_debug.txt"
        debug_file.write_text("\n".join(debug_lines), encoding="utf-8")
        print(f"[INFO] PLIP 파싱 디버그 로그: {debug_file}")

    # 요약 엑셀 (각 complex별 한 줄)
    summary_rows = []
    for base, m in metrics.items():
        summary_rows.append({
            "complex": base,
            "plip_status": statuses.get(base, ""),
            "plip_total_interactions": m["total"],
            "plip_hbond": m["hbond"],
            "plip_hydrophobic": m["hydrophobic"],
            "plip_saltbridge": m["saltbridge"],
        })

    if summary_rows:
        df = pd.DataFrame(summary_rows)
        # 컬럼 순서를 상수로 관리
        df = df[PLIP_SUMMARY_COLS]
        xlsx_path = plip_dir / "plip_summary.xlsx"
        try:
            df.to_excel(xlsx_path, index=False)
            print(f"[INFO] PLIP 요약 엑셀 저장: {xlsx_path}")
        except Exception as e:
            print(f"[WARN] PLIP 요약 엑셀 저장 실패: {e}")

    print(f"[INFO] PLIP 상호작용을 읽어온 구조 수: {len(metrics)}")
    return metrics, statuses


# 각 평가 모델별 점수 범위 (고정 스케일링용)
# 필요하면 나중에 실제 데이터 분포를 보고 조정하면 됨.
PRODIGY_DG_RANGE   = (-20.0, 0.0)   # ΔG: -20 (강한 결합) ~ 0 (거의 결합 없음)
VINA_SCORE_RANGE   = (-15.0, 0.0)   # Vina affinity: -15 ~ 0
PLIP_TOTAL_RANGE   = (0.0, 20.0)    # PLIP total interactions: 0 ~ 20
IPTM_RANGE         = (0.0, 1.0)     # ipTM: 0 ~ 1

def fixed_range_norm(value_dict, vmin, vmax, higher_is_better=True):
    """
    dict(base -> value)를 고정된 [vmin, vmax] 구간 기준으로 0~1로 스케일링.

    - 값이 범위를 벗어나면 vmin/vmax로 clip
    - higher_is_better=True  이면 값이 클수록 1에 가깝게
    - higher_is_better=False 이면 값이 작을수록(에너지가 더 낮을수록) 1에 가깝게
    """
    if vmin == vmax:
        return {}

    # vmin, vmax가 뒤집혀 있으면 정리
    if vmin > vmax:
        vmin, vmax = vmax, vmin

    out = {}
    for k, v in value_dict.items():
        if v is None:
            continue

        # 고정 범위로 클리핑
        x = max(min(v, vmax), vmin)

        if higher_is_better:
            # vmin → 0, vmax → 1
            s = (x - vmin) / (vmax - vmin)
        else:
            # vmin → 1, vmax → 0  (작을수록/더 음수일수록 좋음)
            s = (vmax - x) / (vmax - vmin)

        out[k] = s

    return out


def is_status_ok(status: str) -> bool:
    """
    상태 문자열이 '정상', '정상(txt/log)' 같은 정상 케이스인지 판별.
    """
    if not isinstance(status, str):
        return False
    return status.strip().startswith("정상")


def has_valid_value(v) -> bool:
    """
    값이 None/NaN 이 아닌 실제 숫자인지 판별.
    - 0.0 은 유효한 값으로 인정.
    """
    if v is None:
        return False
    if isinstance(v, (float, int)) and math.isnan(v):
        return False
    return True


def build_and_save_final_table(
    folders,
    peptides,
    rank1_pdbs,
    start_time: datetime | None = None,
    end_time: datetime | None = None,
    step_timings: list[dict] | None = None,
):
    """
    ColabFold / Vina / PLIP / PRODIGY / ipTM 결과를 모아서
    A안 가중치로 FinalScore_A를 계산하고 엑셀로 저장.

    A안:
      PRODIGY 0.50  (ΔG, 더 작을수록 좋음)
      Vina    0.25  (에너지, 더 작을수록 좋음)
      PLIP    0.15  (총 상호작용 수, 많을수록 좋음)
      ipTM    0.10  (인터페이스 신뢰도, 높을수록 좋음)

    추가:
      - 각 평가 모델별 status 컬럼을 엑셀에 함께 기록
        (vina_status, prodigy_status, plip_status)
    """
    results_dir     = folders["results"]
    colabfold_out   = folders["colabfold_out"]
    vina_dir        = folders["vina"]
    plip_dir        = folders["plip"]
    prodigy_dir     = folders["prodigy"]

    # 각 평가 지표 + 상태 불러오기
    vina_vals, vina_status         = load_vina_scores(vina_dir)
    prodigy_vals, prodigy_status   = load_prodigy_scores(prodigy_dir)
    iptm_vals                      = load_iptm_scores(colabfold_out, rank1_pdbs)
    plip_metrics, plip_status      = load_plip_scores(plip_dir)

    # PLIP total 값만 따로 dict로 추출
    plip_total_vals = {b: d.get("total") for b, d in plip_metrics.items()}

    # 고정 범위(물리적 의미 기반) 스케일링
    iptm_norm = fixed_range_norm(
        iptm_vals,
        IPTM_RANGE[0],
        IPTM_RANGE[1],
        higher_is_better=True,   # ipTM은 클수록 좋음
    )
    vina_norm = fixed_range_norm(
        vina_vals,
        VINA_SCORE_RANGE[0],
        VINA_SCORE_RANGE[1],
        higher_is_better=False,  # 에너지는 더 음수일수록 좋음
    )
    prodigy_norm = fixed_range_norm(
        prodigy_vals,
        PRODIGY_DG_RANGE[0],
        PRODIGY_DG_RANGE[1],
        higher_is_better=False,  # ΔG도 더 음수일수록 좋음
    )
    plip_norm = fixed_range_norm(
        plip_total_vals,
        PLIP_TOTAL_RANGE[0],
        PLIP_TOTAL_RANGE[1],
        higher_is_better=True,   # 상호작용 개수는 많을수록 좋음
    )

    # candidate_id → peptide 매핑 (complex_0, complex_1 ...)
    id_to_pep = {f"complex_{i}": pep for i, pep in enumerate(peptides)}

    rows = []
    for pdb_path in rank1_pdbs:
        base = pdb_path.stem                                       # complex_0_unrelaxed_...
        candidate_id = base.split("_unrelaxed")[0]                  # complex_0
        pep_seq = id_to_pep.get(candidate_id, "")

        vina    = vina_vals.get(base)
        prodigy = prodigy_vals.get(base)
        iptm    = iptm_vals.get(base)

        plip_data   = plip_metrics.get(base, {})
        plip_total  = plip_data.get("total")
        plip_hbond  = plip_data.get("hbond")
        plip_hphob  = plip_data.get("hydrophobic")
        plip_salt   = plip_data.get("saltbridge")

        # 체인 수 확인해서 AlphaFold 상태 결정
        chain_counts = get_chain_residue_counts(pdb_path)
        if len(chain_counts) == 1:
            alphafold_status = "신뢰도 있는 결합 복합체 형성 실패(단일체 구조)"
        else:
            alphafold_status = "정상(단백질-펩타이드 복합체)"

        # 이 complex의 status 문자열 가져오기
        vina_st    = vina_status.get(base, "미기록")
        prodigy_st = prodigy_status.get(base, "미기록")
        plip_st    = plip_status.get(base, "미기록")

        # status + 값 유효성 체크
        vina_ok    = is_status_ok(vina_st)    and has_valid_value(vina)
        prodigy_ok = is_status_ok(prodigy_st) and has_valid_value(prodigy)
        plip_ok    = is_status_ok(plip_st)    and has_valid_value(plip_total)

        # 가중치
        w_prodigy = 0.50
        w_vina    = 0.25
        w_plip    = 0.15
        w_iptm    = 0.10

        # 최종 점수 계산 조건:
        #  - 단일체가 아니고
        #  - Vina / PLIP / PRODIGY status가 모두 정상이고
        #  - 해당 값이 실제로 존재할 때만 계산
        if (len(chain_counts) == 1) or not (vina_ok and plip_ok and prodigy_ok):
            final_score = None
        else:
            final_score = (
                w_prodigy * prodigy_norm.get(base, 0.0) +
                w_vina    * vina_norm.get(base, 0.0) +
                w_plip    * plip_norm.get(base, 0.0) +
                w_iptm    * iptm_norm.get(base, 0.0)
            )

        rows.append({
            "candidate_id":     candidate_id,
            "peptide_seq":      pep_seq,
            "complex_pdb":      pdb_path.name,
            "alphafold_status": alphafold_status,
            "final_score":      final_score,
            "prodigy_dG":       prodigy,
            "prodigy_status":   prodigy_st,
            "vina_score":       vina,
            "vina_status":      vina_st,
            "plip_total":       plip_total,
            "plip_hbond":       plip_hbond,
            "plip_hphob":       plip_hphob,
            "plip_salt":        plip_salt,
            "plip_status":      plip_st,
            "iptm":             iptm,
        })


    # FinalScore 기준으로 내림차순 정렬
    # - final_score가 None 인 경우는 가장 아래로 보내고
    # - 나머지는 점수 큰 순서대로 정렬
    rows.sort(
        key=lambda r: (
            r["final_score"] is None,                         # False(0) -> 점수 있음, True(1) -> 점수 없음
            0.0 if r["final_score"] is None else -r["final_score"],  # 점수 있는 것끼리는 -score로 내림차순
        )
    )

    # 엑셀 작성
    wb = Workbook()
    ws = wb.active
    ws.title = "pepbind_ranking"

    headers = FINAL_TABLE_HEADERS
    ws.append(headers)

    # 헤더 이름 → 값 계산 함수 매핑
    value_map = {
        "rank": lambda r, idx: idx,
        "candidate_id": lambda r, idx: r["candidate_id"],
        "peptide_seq": lambda r, idx: r["peptide_seq"],
        "complex_pdb": lambda r, idx: r["complex_pdb"],
        "AlphaFold_status": lambda r, idx: r["alphafold_status"],
        "FinalScore": lambda r, idx: round(r["final_score"], 4) if r["final_score"] is not None else None,
        "PRODIGY_dG(kcal/mol)": lambda r, idx: r["prodigy_dG"],
        "PRODIGY_status": lambda r, idx: r["prodigy_status"],
        "Vina_score(kcal/mol)": lambda r, idx: r["vina_score"],
        "Vina_status": lambda r, idx: r["vina_status"],
        "PLIP_total_interactions": lambda r, idx: r["plip_total"],
        "PLIP_hbond": lambda r, idx: r["plip_hbond"],
        "PLIP_hydrophobic": lambda r, idx: r["plip_hphob"],
        "PLIP_saltbridge": lambda r, idx: r["plip_salt"],
        "PLIP_status": lambda r, idx: r["plip_status"],
        "ipTM": lambda r, idx: r["iptm"],
    }

    for idx, r in enumerate(rows, start=1):
        row_vals = [value_map[h](r, idx) for h in headers]
        ws.append(row_vals)
    
    # dG와 같은 실수 컬럼에 소수점 2자리 표시 형식 적용
    for col_idx, header in enumerate(headers, start=1):
        if "dG" in header:  # "PRODIGY_dG(kcal/mol)" 컬럼
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"  # 항상 소수점 둘째 자리까지


    # ─────────────────────────────────────────────
    # 두 번째 시트: setting
    # ─────────────────────────────────────────────
    ws_setting = wb.create_sheet(title="setting")

    ws_setting.append(["항목", "값"])

    target_seq_str = TARGET_SEQUENCE.strip()
    target_len = len(TARGET_SEQUENCE.strip())

    ws_setting.append(["TARGET_SEQUENCE", target_seq_str])
    ws_setting.append(["TARGET_SEQUENCE_length", target_len])
    ws_setting.append(["PEPTIDE_LENGTH", PEPTIDE_LENGTH])
    ws_setting.append(["NUM_PEPTIDES", NUM_PEPTIDES])
    ws_setting.append(["PepMLM_temperature", PEPMLM_TEMPERATURE])
    ws_setting.append(["PepMLM_top_k", PEPMLM_TOP_K])

    # 코드 실행 시작/종료/총 소요시간
    if start_time is not None:
        ws_setting.append(
            ["code_start_time", start_time.strftime("%Y.%m.%d %H:%M:%S")]
        )
    if end_time is not None:
        ws_setting.append(
            ["code_end_time", end_time.strftime("%Y.%m.%d %H:%M:%S")]
        )
    if start_time is not None and end_time is not None:
        ws_setting.append(
            [
                "code_total_elapsed",
                format_elapsed(start_time, end_time),
            ]
        )

        # 후보군(샘플) 1개당 평균 소요 시간 (시간/분/초 + 초 표시)
        n_samples = len(peptides) if peptides is not None else 0
        if n_samples > 0:
            total_seconds = (end_time - start_time).total_seconds()
            per_sample_sec = total_seconds / n_samples
            per_sample_hms = format_seconds_hms(per_sample_sec)

            ws_setting.append(
                [
                    "code_elapsed_per_sample",
                    f"{per_sample_hms} (≈ {per_sample_sec:.2f}초/샘플)",
                ]
            )

    # 빈 줄 하나
    ws_setting.append([])
    # 스텝별 시간 테이블 헤더
    ws_setting.append(["step_label", "start_time", "end_time", "elapsed"])

    if step_timings:
        for rec in step_timings:
            ws_setting.append(
                [
                    rec["step"],
                    rec["start"].strftime("%Y.%m.%d %H:%M:%S"),
                    rec["end"].strftime("%Y.%m.%d %H:%M:%S"),
                    rec["elapsed"],
                ]
            )

    # 열 너비 자동조절 (랭킹 시트 + setting 시트 둘 다)
    autofit_worksheet_columns(ws)
    autofit_worksheet_columns(ws_setting)

    # '값' 열은 너무 넓어지지 않도록 고정 너비(예: 20)로 설정
    for cell in ws_setting[1]:  # 첫 번째 행 헤더에서 '값'이 있는 열 찾기
        if cell.value == "값":
            col_letter = get_column_letter(cell.column)
            ws_setting.column_dimensions[col_letter].width = 20
            break

    out_xlsx = results_dir / f"final_peptide_rank_{timestamp()}.xlsx"
    wb.save(out_xlsx)
    print(f"✅ 최종 결과 엑셀 저장: {out_xlsx}")
    return out_xlsx


# =====================================================================
# === MAIN ============================================================
# =====================================================================

def main():
    # 전체 파이프라인 시작 시간
    global START_TIME
    START_TIME = datetime.now()

    # STEP 1: 워크스페이스/폴더 구조 생성
    step1_start = datetime.now()
    folders = init_workspace()
    step1_end = datetime.now()
    print_step_timing("STEP 1: 워크스페이스 / 폴더 구조 생성", step1_start, step1_end)

    # ★ 워크스페이스 안에 로그 파일 생성 + stdout/stderr를 동시에 기록
    log_path = folders["results"] / f"pepbind_run_{timestamp()}.log"
    setup_logging(log_path)

    # ★ 이번 실행에 사용된 옵션/환경 요약 출력 (→ 자동으로 로그에도 남음)
    print_run_config()

    # STEP 2: 타깃 FASTA + PepMLM 기반 펩타이드 생성
    step2_start = datetime.now()

    target_seq = TARGET_SEQUENCE.strip()
    target_fasta = write_target_fasta(folders["fasta"], target_seq)
    print(f"✔️ 타깃 단백질 길이: {len(target_seq)}")
    print(f"✔️ 타깃 FASTA: {target_fasta}")

    tokenizer, model = load_esm_mlm()
    peptides = generate_peptides_with_mlm(
        tokenizer,
        model,
        target_seq,
        num_peptides=NUM_PEPTIDES,
        peptide_len=PEPTIDE_LENGTH,
        top_k=PEPMLM_TOP_K,
        temperature=PEPMLM_TEMPERATURE,
    )
    pep_fasta = write_peptide_fasta(folders["fasta"], peptides)
    print(f"✔️ PepMLM 결과 저장: {pep_fasta}")

    # ESM 모델은 이후 사용하지 않으므로 메모리에서 제거 + GPU 캐시 정리
    try:
        del model
        del tokenizer
    except NameError:
        pass
    clear_gpu_memory()

    step2_end = datetime.now()
    print_step_timing("STEP 2: PepMLM 기반 펩타이드 생성", step2_start, step2_end)

    # STEP 3: ColabFold 구조 예측
    rank1_pdbs = []
    if RUN_COLABFOLD and peptides:
        step3_start = datetime.now()
        csv_path = prepare_colabfold_batch_csv(
            folders["temp"],
            target_seq,
            peptides,
        )
        try:
            rank1_pdbs = run_colabfold_batch_with_progress(
                csv_path,
                folders["colabfold_out"],
                total_complexes=len(peptides),
            )
        except RuntimeError as e:
            step3_end = datetime.now()
            print_step_timing("STEP 3: ColabFold 구조 예측 (실패)", step3_start, step3_end)
            print("\n[ERROR] ColabFold 단계에서 오류가 발생하여 파이프라인을 중단합니다.")
            print("       메시지:", e)
            return
        step3_end = datetime.now()
        print_step_timing("STEP 3: ColabFold 구조 예측", step3_start, step3_end)
    else:
        now = datetime.now()
        print("\n[INFO] RUN_COLABFOLD=False 또는 펩타이드 없음 → ColabFold 단계 스킵")
        print_step_timing("STEP 3: ColabFold 구조 예측 (스킵)", now, now)

    # STEP 4: Vina
    if RUN_VINA:
        step4_start = datetime.now()
        run_vina_on_rank1(rank1_pdbs, folders["vina"])
        step4_end = datetime.now()
        print_step_timing("STEP 4: AutoDock Vina 도킹", step4_start, step4_end)
    else:
        now = datetime.now()
        print("\n[INFO] RUN_VINA=False → Vina 단계 스킵")
        print_step_timing("STEP 4: AutoDock Vina 도킹 (스킵)", now, now)

    # STEP 5: PLIP
    if RUN_PLIP:
        step5_start = datetime.now()
        run_plip_on_rank1(rank1_pdbs, folders["plip"])
        step5_end = datetime.now()
        print_step_timing("STEP 5: PLIP 상호작용 분석", step5_start, step5_end)
    else:
        now = datetime.now()
        print("\n[INFO] RUN_PLIP=False → PLIP 단계 스킵")
        print_step_timing("STEP 5: PLIP 상호작용 분석 (스킵)", now, now)

    # STEP 6: PRODIGY
    if RUN_PRODIGY:
        step6_start = datetime.now()
        run_prodigy_on_rank1(rank1_pdbs, folders["prodigy"])
        step6_end = datetime.now()
        print_step_timing("STEP 6: PRODIGY 결합 친화도 평가", step6_start, step6_end)
    else:
        now = datetime.now()
        print("\n[INFO] RUN_PRODIGY=False → PRODIGY 단계 스킵")
        print_step_timing("STEP 6: PRODIGY 결합 친화도 평가 (스킵)", now, now)

    # STEP 7: rank_001 PDB zip + 최종 엑셀
    step7_start = datetime.now()
    pdb_zip = None
    final_xlsx = None

    if rank1_pdbs:
        pdb_zip = zip_rank1_pdbs(rank1_pdbs, folders["results"])
    else:
        print("[INFO] rank_001 PDB가 없어 zip/엑셀 생성을 생략합니다.")

    step7_end = datetime.now()
    print_step_timing("STEP 7: 결과 zip / 최종 엑셀 생성", step7_start, step7_end)

    # 전체 파이프라인 종료 시간 및 소요 시간
    global END_TIME
    END_TIME = datetime.now()

    # STEP_TIMINGS / START_TIME / END_TIME 을 포함해서 최종 엑셀 생성
    if rank1_pdbs:
        final_xlsx = build_and_save_final_table(
            folders,
            peptides,
            rank1_pdbs,
            start_time=START_TIME,
            end_time=END_TIME,
            step_timings=STEP_TIMINGS,
        )

    print("\n" + "=" * 80)
    print("🎉 파이프라인 실행 종료")
    print("=" * 80)
    print(f"[INFO] 워크스페이스: {folders['root']}")
    if pdb_zip:
        print(f"[INFO] PDB zip: {pdb_zip}")
    if final_xlsx:
        print(f"[INFO] 최종 엑셀: {final_xlsx}")
    print(f"[INFO] 시작 시간: {START_TIME.strftime('%Y.%m.%d %H:%M:%S')}")
    print(f"[INFO] 종료 시간: {END_TIME.strftime('%Y.%m.%d %H:%M:%S')}")
    print(f"[INFO] 총 소요 시간: {format_elapsed(START_TIME, END_TIME)}")
    print("=" * 80)


if __name__ == "__main__":
    main()
